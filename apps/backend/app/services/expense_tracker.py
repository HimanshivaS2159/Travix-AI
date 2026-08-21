"""
Trip Expense Tracker Service
Handles Excel-based expense tracking with CRUD operations
"""

import os
from datetime import datetime
from typing import List, Optional, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

# Excel file path
EXCEL_FILE_PATH = "data/trip_expenses.xlsx"
SHEET_NAME = "Expenses"

# Column headers
HEADERS = ["ID", "Trip Name", "Date", "Category", "Amount", "Notes"]
CATEGORIES = ["Food", "Travel", "Stay", "Shopping", "Other"]


class ExpenseTracker:
    """Excel-based expense tracker"""

    def __init__(self):
        self.file_path = EXCEL_FILE_PATH
        self._ensure_excel_file()

    def _ensure_excel_file(self):
        """Create Excel file with headers if it doesn't exist"""
        # Ensure data directory exists
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        if not os.path.exists(self.file_path):
            logger.info(f"Creating new Excel file: {self.file_path}")
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME

            # Write headers with styling
            for col, header in enumerate(HEADERS, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set column widths
            ws.column_dimensions['A'].width = 8   # ID
            ws.column_dimensions['B'].width = 25  # Trip Name
            ws.column_dimensions['C'].width = 12  # Date
            ws.column_dimensions['D'].width = 12  # Category
            ws.column_dimensions['E'].width = 12  # Amount
            ws.column_dimensions['F'].width = 40  # Notes

            wb.save(self.file_path)
            logger.info("Excel file created successfully")

    def _get_next_id(self) -> int:
        """Get the next available ID"""
        wb = load_workbook(self.file_path)
        ws = wb[SHEET_NAME]
        
        if ws.max_row == 1:  # Only headers
            return 1
        
        # Get the last ID
        last_id = ws.cell(row=ws.max_row, column=1).value
        return (last_id or 0) + 1

    def add_expense(
        self,
        trip_name: str,
        date: str,
        category: str,
        amount: float,
        notes: Optional[str] = ""
    ) -> Dict[str, Any]:
        """
        Add a new expense entry to Excel
        
        Args:
            trip_name: Name of the trip
            date: Date in YYYY-MM-DD format
            category: Expense category (Food, Travel, Stay, Shopping, Other)
            amount: Expense amount
            notes: Optional notes
            
        Returns:
            Dict with the created expense
        """
        try:
            # Validate category
            if category not in CATEGORIES:
                raise ValueError(f"Invalid category. Must be one of: {', '.join(CATEGORIES)}")

            # Load workbook
            wb = load_workbook(self.file_path)
            ws = wb[SHEET_NAME]

            # Get next ID
            expense_id = self._get_next_id()

            # Append new row
            new_row = [expense_id, trip_name, date, category, amount, notes or ""]
            ws.append(new_row)

            # Format the new row
            row_num = ws.max_row
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Format amount as currency
                if col == 5:  # Amount column
                    cell.number_format = '₹#,##0.00'

            # Save workbook
            wb.save(self.file_path)
            logger.info(f"Added expense ID {expense_id} to Excel")

            return {
                "id": expense_id,
                "trip_name": trip_name,
                "date": date,
                "category": category,
                "amount": amount,
                "notes": notes or ""
            }

        except Exception as e:
            logger.error(f"Error adding expense: {e}")
            raise

    def get_all_expenses(self) -> List[Dict[str, Any]]:
        """
        Get all expenses from Excel
        
        Returns:
            List of expense dictionaries
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[SHEET_NAME]

            expenses = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] is not None:  # Check if row has data
                    expenses.append({
                        "id": row[0],
                        "trip_name": row[1],
                        "date": row[2] if isinstance(row[2], str) else row[2].strftime("%Y-%m-%d") if row[2] else "",
                        "category": row[3],
                        "amount": float(row[4]) if row[4] else 0.0,
                        "notes": row[5] or ""
                    })

            return expenses

        except Exception as e:
            logger.error(f"Error getting expenses: {e}")
            raise

    def get_expense_by_id(self, expense_id: int) -> Optional[Dict[str, Any]]:
        """
        Get a specific expense by ID
        
        Args:
            expense_id: Expense ID
            
        Returns:
            Expense dict or None if not found
        """
        expenses = self.get_all_expenses()
        for expense in expenses:
            if expense["id"] == expense_id:
                return expense
        return None

    def update_expense(
        self,
        expense_id: int,
        trip_name: Optional[str] = None,
        date: Optional[str] = None,
        category: Optional[str] = None,
        amount: Optional[float] = None,
        notes: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Update an existing expense
        
        Args:
            expense_id: ID of expense to update
            trip_name: New trip name (optional)
            date: New date (optional)
            category: New category (optional)
            amount: New amount (optional)
            notes: New notes (optional)
            
        Returns:
            Updated expense dict
        """
        try:
            # Validate category if provided
            if category and category not in CATEGORIES:
                raise ValueError(f"Invalid category. Must be one of: {', '.join(CATEGORIES)}")

            wb = load_workbook(self.file_path)
            ws = wb[SHEET_NAME]

            # Find the row with matching ID
            row_num = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == expense_id:
                    row_num = row
                    break

            if row_num is None:
                raise ValueError(f"Expense with ID {expense_id} not found")

            # Update fields
            if trip_name is not None:
                ws.cell(row=row_num, column=2, value=trip_name)
            if date is not None:
                ws.cell(row=row_num, column=3, value=date)
            if category is not None:
                ws.cell(row=row_num, column=4, value=category)
            if amount is not None:
                cell = ws.cell(row=row_num, column=5, value=amount)
                cell.number_format = '₹#,##0.00'
            if notes is not None:
                ws.cell(row=row_num, column=6, value=notes)

            # Save workbook
            wb.save(self.file_path)
            logger.info(f"Updated expense ID {expense_id}")

            # Return updated expense
            return self.get_expense_by_id(expense_id)

        except Exception as e:
            logger.error(f"Error updating expense: {e}")
            raise

    def delete_expense(self, expense_id: int) -> bool:
        """
        Delete an expense by ID
        
        Args:
            expense_id: ID of expense to delete
            
        Returns:
            True if deleted, False if not found
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[SHEET_NAME]

            # Find the row with matching ID
            row_num = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == expense_id:
                    row_num = row
                    break

            if row_num is None:
                return False

            # Delete the row
            ws.delete_rows(row_num)

            # Save workbook
            wb.save(self.file_path)
            logger.info(f"Deleted expense ID {expense_id}")

            return True

        except Exception as e:
            logger.error(f"Error deleting expense: {e}")
            raise

    def get_expenses_by_trip(self, trip_name: str) -> List[Dict[str, Any]]:
        """Get all expenses for a specific trip"""
        all_expenses = self.get_all_expenses()
        return [e for e in all_expenses if e["trip_name"].lower() == trip_name.lower()]

    def get_expenses_by_category(self, category: str) -> List[Dict[str, Any]]:
        """Get all expenses for a specific category"""
        all_expenses = self.get_all_expenses()
        return [e for e in all_expenses if e["category"] == category]

    def get_expenses_by_date_range(self, start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """Get expenses within a date range (YYYY-MM-DD format)"""
        all_expenses = self.get_all_expenses()
        filtered = []
        
        for expense in all_expenses:
            expense_date = expense["date"]
            if start_date <= expense_date <= end_date:
                filtered.append(expense)
        
        return filtered

    def get_trip_summary(self) -> Dict[str, Any]:
        """
        Get summary of expenses grouped by trip
        
        Returns:
            Dict with trip summaries and grand total
        """
        expenses = self.get_all_expenses()
        
        trip_totals = {}
        category_totals = {}
        grand_total = 0.0

        for expense in expenses:
            trip_name = expense["trip_name"]
            category = expense["category"]
            amount = expense["amount"]

            # Trip totals
            if trip_name not in trip_totals:
                trip_totals[trip_name] = {
                    "total": 0.0,
                    "count": 0,
                    "expenses": []
                }
            trip_totals[trip_name]["total"] += amount
            trip_totals[trip_name]["count"] += 1
            trip_totals[trip_name]["expenses"].append(expense)

            # Category totals
            if category not in category_totals:
                category_totals[category] = 0.0
            category_totals[category] += amount

            # Grand total
            grand_total += amount

        return {
            "trips": trip_totals,
            "categories": category_totals,
            "grand_total": grand_total,
            "total_expenses": len(expenses)
        }

    def get_all_trip_names(self) -> List[str]:
        """Get unique list of all trip names"""
        expenses = self.get_all_expenses()
        trip_names = list(set([e["trip_name"] for e in expenses]))
        return sorted(trip_names)


# Global instance
expense_tracker = ExpenseTracker()
